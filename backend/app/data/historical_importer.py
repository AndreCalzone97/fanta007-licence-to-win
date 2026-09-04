from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from app.domain.player import Player, PlayerDataset, PlayerSeasonStats, StatsConfidence


COLUMN_ALIASES = {
    "id": ["id"],
    "name": ["calciatore", "giocatore", "nome", "player"],
    "team": ["sq", "squadra", "team"],
    "competition": ["nazione", "competizione", "competition"],
    "appearances": ["pv", "presenze", "partite"],
    "average_rating": ["mv", "media voto", "media_voto"],
    "fantasy_average": ["fm", "fantamedia", "media fantavoto"],
    "goals": ["gol", "goal", "gf"],
    "goals_conceded": ["gs", "gol subiti"],
    "assists": ["ass", "assist"],
    "yellow_cards": ["amm", "ammonizioni"],
    "red_cards": ["esp", "espulsioni"],
    "penalties_saved": ["rp", "rigori parati"],
    "penalties_taken": ["rc", "rigori calciati", "rigori tirati"],
    "penalties_scored": ["r+", "rigori segnati"],
    "penalties_missed": ["r-", "rigori sbagliati"],
    "own_goals": ["au", "aut", "autogol"],
}


@dataclass(frozen=True)
class ImportedPlayerStats:
    player_id: int
    player_name: str
    role: str
    stats: PlayerSeasonStats


def _normalize_header(value: str) -> str:
    return " ".join(value.strip().casefold().replace("_", " ").split())


def _find(row: dict[str, object], field: str):
    normalized = {_normalize_header(key): value for key, value in row.items()}
    return next((normalized[alias] for alias in COLUMN_ALIASES[field] if alias in normalized), None)


def _number(value: object | None, integer: bool = True):
    if value is None:
        return None
    cleaned = str(value).strip()
    if not cleaned or cleaned.casefold() in {"-", "n/d", "nan", "none"}:
        return None
    parsed = float(cleaned.replace(".", "").replace(",", ".")) if "," in cleaned else float(cleaned)
    return int(parsed) if integer else parsed


def _competition_name(value: object | None, fallback: str) -> str:
    if value is None or not str(value).strip():
        return fallback
    normalized = str(value).strip()
    aliases = {"liga": "La Liga", "liga1": "Ligue 1", "serie a": "Serie A"}
    return aliases.get(normalized.casefold(), normalized)


def read_fantacalcio_xlsx(
    path: Path,
    season: str,
    source_url: str,
    *,
    source: str = "Fantacalcio.it",
    default_competition: str = "Serie A",
    confidence: StatsConfidence = "HIGH",
    updated_at: date | None = None,
) -> list[ImportedPlayerStats]:
    """Read an official Fantacalcio export and preserve its numeric ID as identity key."""

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    worksheet = workbook["Tutti"] if "Tutti" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    next(rows, None)  # Human-readable export title.
    raw_headers = next(rows, None)
    if not raw_headers:
        return []
    headers = [_normalize_header(str(value)) if value is not None else "" for value in raw_headers]
    imported: list[ImportedPlayerStats] = []
    seen: dict[int, ImportedPlayerStats] = {}

    for values in rows:
        row = {header: value for header, value in zip(headers, values) if header}
        player_id = _number(_find(row, "id"))
        name = _find(row, "name")
        role = _find(row, "role") if "role" in COLUMN_ALIASES else row.get("r")
        if player_id is None or name is None:
            continue
        appearances = _number(_find(row, "appearances"))
        average_rating = _number(_find(row, "average_rating"), False)
        fantasy_average = _number(_find(row, "fantasy_average"), False)
        # In official exports 0/0 with PV=0 means no voteable sample, not a real rating.
        if appearances == 0:
            average_rating = None if average_rating == 0 else average_rating
            fantasy_average = None if fantasy_average == 0 else fantasy_average
        team = _find(row, "team")
        stats = PlayerSeasonStats(
            season=season,
            competition=_competition_name(_find(row, "competition"), default_competition),
            club=str(team).strip() if team is not None and str(team).strip() else None,
            source=source,
            source_url=source_url,
            updated_at=updated_at,
            confidence=confidence,
            appearances=appearances,
            average_rating=average_rating,
            fantasy_average=fantasy_average,
            goals=_number(_find(row, "goals")),
            goals_conceded=_number(_find(row, "goals_conceded")),
            assists=_number(_find(row, "assists")),
            yellow_cards=_number(_find(row, "yellow_cards")),
            red_cards=_number(_find(row, "red_cards")),
            penalties_taken=_number(_find(row, "penalties_taken")),
            penalties_scored=_number(_find(row, "penalties_scored")),
            penalties_missed=_number(_find(row, "penalties_missed")),
            penalties_saved=_number(_find(row, "penalties_saved")),
            own_goals=_number(_find(row, "own_goals")),
        )
        record = ImportedPlayerStats(int(player_id), str(name).strip(), str(role or "").strip(), stats)
        previous = seen.get(record.player_id)
        if previous is None or (record.stats.appearances or 0) > (previous.stats.appearances or 0):
            seen[record.player_id] = record

    imported.extend(seen.values())
    workbook.close()
    return imported


def stats_source_priority(source: str) -> int:
    normalized = source.casefold()
    if "euroleghe" in normalized:
        return 200
    if "fantacalcio" in normalized:
        return 300
    return 100


def merge_player_statistics(
    dataset: PlayerDataset,
    sources: Iterable[tuple[int, Iterable[ImportedPlayerStats]]],
) -> PlayerDataset:
    """Merge one authoritative record per player/season, preferring source priority then sample."""

    player_ids = {player.id for player in dataset.players}
    selected: dict[tuple[int, str], tuple[int, PlayerSeasonStats]] = {}
    for player in dataset.players:
        for stats in player.statistics:
            key = (player.id, stats.season)
            selected[key] = (stats_source_priority(stats.source), stats)

    def completeness(stats: PlayerSeasonStats) -> int:
        fields = (
            stats.appearances, stats.average_rating, stats.fantasy_average, stats.goals,
            stats.goals_conceded, stats.assists, stats.yellow_cards, stats.red_cards,
            stats.penalties_taken, stats.penalties_scored, stats.penalties_missed, stats.penalties_saved, stats.own_goals,
        )
        return sum(value is not None for value in fields)

    for priority, rows in sources:
        for row in rows:
            if row.player_id not in player_ids:
                continue
            key = (row.player_id, row.stats.season)
            previous = selected.get(key)
            incoming_appearances = row.stats.appearances or 0
            previous_appearances = previous[1].appearances or 0 if previous else -1
            if previous is None or priority > previous[0] or (
                priority == previous[0]
                and (
                    incoming_appearances > previous_appearances
                    or (incoming_appearances == previous_appearances and completeness(row.stats) > completeness(previous[1]))
                )
            ):
                selected[key] = (priority, row.stats)

    def season_key(stats: PlayerSeasonStats) -> tuple[int, str]:
        start = int(stats.season[:4]) if stats.season[:4].isdigit() else 0
        return start, stats.competition

    updated_players: list[Player] = []
    for player in dataset.players:
        seasons = [stats for (player_id, _), (_, stats) in selected.items() if player_id == player.id]
        updated_players.append(player.model_copy(update={"statistics": sorted(seasons, key=season_key, reverse=True)}))
    return dataset.model_copy(update={"players": updated_players})


def read_historical_csv(path: Path, season: str, source_url: str) -> list[tuple[str, str | None, PlayerSeasonStats]]:
    with Path(path).open(encoding="utf-8-sig", newline="") as stream:
        sample = stream.read(4096)
        stream.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        rows = csv.DictReader(stream, dialect=dialect)
        imported = []
        for row in rows:
            name = _find(row, "name")
            if not name:
                continue
            imported.append((name.strip(), (_find(row, "team") or "").strip() or None, PlayerSeasonStats(
                season=season, competition="Serie A", source="Fantacalcio.it", source_url=source_url,
                appearances=_number(_find(row, "appearances")), average_rating=_number(_find(row, "average_rating"), False),
                fantasy_average=_number(_find(row, "fantasy_average"), False), goals=_number(_find(row, "goals")),
                goals_conceded=_number(_find(row, "goals_conceded")), assists=_number(_find(row, "assists")),
                yellow_cards=_number(_find(row, "yellow_cards")), red_cards=_number(_find(row, "red_cards")),
                penalties_taken=_number(_find(row, "penalties_taken")), penalties_saved=_number(_find(row, "penalties_saved")), own_goals=_number(_find(row, "own_goals")),
            )))
    return imported
