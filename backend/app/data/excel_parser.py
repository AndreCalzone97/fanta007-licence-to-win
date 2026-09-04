from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_HEADERS = (
    "Id",
    "R",
    "RM",
    "Nome",
    "Squadra",
    "Qt.A",
    "Qt.I",
    "Diff.",
    "Qt.A M",
    "Qt.I M",
    "Diff.M",
    "FVM",
    "FVM M",
)


class DatasetValidationError(ValueError):
    """Raised when the source workbook does not match the expected contract."""


@dataclass(frozen=True, slots=True)
class RawPlayerRow:
    id: int
    role_classic: str
    role_mantra_raw: str
    source_name: str
    team: str
    current_quotation: int
    initial_quotation: int
    quotation_delta: int
    current_quotation_mantra: int
    initial_quotation_mantra: int
    quotation_delta_mantra: int
    fvm: int
    fvm_mantra: int


def _required_text(value: Any, field: str, excel_row: int) -> str:
    if not isinstance(value, str) or not value.strip():
        raise DatasetValidationError(
            f"Riga {excel_row}: {field} deve essere una stringa non vuota."
        )
    return value.strip()


def _required_int(value: Any, field: str, excel_row: int) -> int:
    if isinstance(value, bool) or not isinstance(value, int):
        raise DatasetValidationError(
            f"Riga {excel_row}: {field} deve essere un intero, ricevuto {value!r}."
        )
    return value


def parse_players(source: Path, sheet_name: str = "Tutti") -> list[RawPlayerRow]:
    source = Path(source)
    if not source.is_file():
        raise FileNotFoundError(f"Workbook non trovato: {source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise DatasetValidationError(
                f"Foglio {sheet_name!r} assente. Disponibili: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        headers = tuple(cell.value for cell in worksheet[2])
        if headers != EXPECTED_HEADERS:
            raise DatasetValidationError(
                f"Header non valido in {sheet_name}: {headers!r}"
            )

        players: list[RawPlayerRow] = []
        seen_ids: set[int] = set()

        for excel_row, values in enumerate(
            worksheet.iter_rows(min_row=3, max_col=len(EXPECTED_HEADERS), values_only=True),
            start=3,
        ):
            if all(value is None for value in values):
                continue

            player = RawPlayerRow(
                id=_required_int(values[0], "Id", excel_row),
                role_classic=_required_text(values[1], "R", excel_row),
                role_mantra_raw=_required_text(values[2], "RM", excel_row),
                source_name=_required_text(values[3], "Nome", excel_row),
                team=_required_text(values[4], "Squadra", excel_row),
                current_quotation=_required_int(values[5], "Qt.A", excel_row),
                initial_quotation=_required_int(values[6], "Qt.I", excel_row),
                quotation_delta=_required_int(values[7], "Diff.", excel_row),
                current_quotation_mantra=_required_int(values[8], "Qt.A M", excel_row),
                initial_quotation_mantra=_required_int(values[9], "Qt.I M", excel_row),
                quotation_delta_mantra=_required_int(values[10], "Diff.M", excel_row),
                fvm=_required_int(values[11], "FVM", excel_row),
                fvm_mantra=_required_int(values[12], "FVM M", excel_row),
            )

            if player.id in seen_ids:
                raise DatasetValidationError(f"Id duplicato: {player.id}")
            seen_ids.add(player.id)

            if player.role_classic not in {"P", "D", "C", "A"}:
                raise DatasetValidationError(
                    f"Riga {excel_row}: ruolo Classic non valido {player.role_classic!r}."
                )
            if player.quotation_delta != (
                player.current_quotation - player.initial_quotation
            ):
                raise DatasetValidationError(
                    f"Riga {excel_row}: Diff. incoerente per {player.source_name}."
                )
            if player.quotation_delta_mantra != (
                player.current_quotation_mantra - player.initial_quotation_mantra
            ):
                raise DatasetValidationError(
                    f"Riga {excel_row}: Diff.M incoerente per {player.source_name}."
                )

            players.append(player)

        if not players:
            raise DatasetValidationError(f"Il foglio {sheet_name!r} non contiene giocatori.")
        return players
    finally:
        workbook.close()

