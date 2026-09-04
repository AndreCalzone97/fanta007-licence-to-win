from pathlib import Path
from openpyxl import load_workbook

from app.domain.player import PlayerDataset

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATASET = PROJECT_ROOT / "data" / "normalized" / "players.json"
SOURCE_DIR = PROJECT_ROOT / "data" / "source"

FIELD_MAP = {
    "Pv": "appearances",
    "Mv": "average_rating",
    "Fm": "fantasy_average",
    "Gf": "goals",
    "Gs": "goals_conceded",
    "Rp": "penalties_saved",
    "Rc": "penalties_taken",
    "R+": "penalties_scored",
    "R-": "penalties_missed",
    "Ass": "assists",
    "Amm": "yellow_cards",
    "Esp": "red_cards",
    "Au": "own_goals",
}


def _rows(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Tutti"]
    iterator = ws.iter_rows(values_only=True)
    next(iterator)
    headers = list(next(iterator))
    selected = {}
    for values in iterator:
        row = dict(zip(headers, values))
        if row.get("Id") is None:
            continue
        player_id = int(row["Id"])
        if player_id not in selected or (row.get("Pv") or 0) > (selected[player_id].get("Pv") or 0):
            selected[player_id] = row
    wb.close()
    return selected


def _expected_value(row, column):
    value = row.get(column)
    if column in {"Mv", "Fm"} and (row.get("Pv") or 0) == 0 and value == 0:
        return None
    return value


def test_active_dataset_matches_official_excel_stats_field_by_field():
    dataset = PlayerDataset.model_validate_json(DATASET.read_text(encoding="utf-8"))
    active = {(p.id, s.season): s for p in dataset.players for s in p.statistics}
    player_ids = {p.id for p in dataset.players}

    serie_a_2026 = _rows(SOURCE_DIR / "Statistiche_Fantacalcio_Stagione_2026_27.xlsx")
    serie_a_2025 = _rows(SOURCE_DIR / "Statistiche_Fantacalcio_Stagione_2025_26.xlsx")
    euro_2025 = _rows(SOURCE_DIR / "Statistiche_Fantacalcio_EuroLeghe_Stagione_2025_26.xlsx")

    expected = {}
    for player_id in player_ids:
        if player_id in serie_a_2026:
            expected[(player_id, "2026/27")] = serie_a_2026[player_id]
        if player_id in serie_a_2025:
            expected[(player_id, "2025/26")] = serie_a_2025[player_id]
        elif player_id in euro_2025:
            expected[(player_id, "2025/26")] = euro_2025[player_id]

    assert len(expected) == 921
    assert set(active) == set(expected)

    for key, row in expected.items():
        season = active[key]
        for excel_column, model_field in FIELD_MAP.items():
            assert getattr(season, model_field) == _expected_value(row, excel_column), (key, excel_column)
